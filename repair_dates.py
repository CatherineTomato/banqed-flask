"""Repair the sale dates that the first migration failed to parse.
 
The old database stored dates as datetime strings like '2026-07-09 00:00:00',
a format the migration's parser did not recognise, so date_listed and
date_sold were left blank on every item. The spreadsheet still holds the
original values, so this reads them back from there.
 
Run this once, after migrate_merge_sales.py:
 
    python3 repair_dates.py
 
Safe to run more than once — it only fills in dates that are currently blank
unless you pass --force.
"""
 
import sys
from datetime import date, datetime
 
import openpyxl
 
from app import app
from models import db, Item
 
WORKBOOK = "data/BANQED.xlsx"
 
DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",   # the format the old database used
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d-%b-%Y",
    "%d %b %Y",
    "%m/%d/%Y",
]
 
 
def parse_date(value):
    """Accept real dates, datetimes, and the several text shapes in the sheet."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
 
    text = str(value).strip()
    if not text:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    print(f"  ! still could not parse {text!r}")
    return None
 
 
def main(force=False):
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = workbook["Sales"]
 
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    index = {h: i for i, h in enumerate(headers) if h}
 
    listed = sold = skipped = unmatched = 0
 
    with app.app_context():
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[index["Item name"]]
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
 
            item = Item.query.filter(
                db.func.lower(Item.item_name) == name.lower()
            ).first()
 
            if item is None:
                unmatched += 1
                continue
 
            new_listed = parse_date(row[index["Date Listed"]])
            new_sold = parse_date(row[index["Date Sold"]])
 
            if new_listed and (force or item.date_listed is None):
                item.date_listed = new_listed
                listed += 1
            elif new_listed:
                skipped += 1
 
            if new_sold and (force or item.date_sold is None):
                item.date_sold = new_sold
                sold += 1
 
        db.session.commit()
 
        print(f"Restored {listed} listing dates and {sold} sale dates.")
        if skipped:
            print(f"Left {skipped} existing dates alone (use --force to overwrite).")
        if unmatched:
            print(f"{unmatched} sheet rows had no matching item.")
 
        with_days = [
            i for i in Item.query.all()
            if i.days_to_sell is not None
        ]
        print(f"\nDays to sell now calculates on {len(with_days)} items.")
        for item in with_days[:5]:
            print(f"  {item.item_name}: listed {item.date_listed}, "
                  f"sold {item.date_sold}, {item.days_to_sell} days")
 
 
if __name__ == "__main__":
    main(force="--force" in sys.argv)