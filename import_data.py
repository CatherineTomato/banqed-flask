"""Import the BANQED spreadsheet into the database.
 
Both sheets feed the same table now: Sales rows are wardrobe items that
happen to be for sale, matched back to their Wardrobe row by name. This
mirrors how the spreadsheet actually worked — every Sales row there is a
Wardrobe item, never a separate thing.
 
    python3 import_data.py
"""
 
from datetime import date, datetime
 
import openpyxl
 
from app import app
from models import db, Item, SALES_RESALE
 
WORKBOOK = "data/BANQED.xlsx"
 
DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",   # datetime strings, e.g. 2026-07-09 00:00:00
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d-%b-%Y",
    "%d %b %Y",
    "%m/%d/%Y",
]
 
 
def clean(value):
    if value is None:
        return ""
    return str(value).strip()
 
 
def clean_number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
 
 
def clean_date(value):
    """The sheet mixes real dates with text, so accept both."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
 
    text = str(value).strip()
    if not text:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
 
 
def sheet_rows(workbook, sheet_name):
    """Yield each row of a sheet as a {header: value} dict."""
    ws = workbook[sheet_name]
    headers = [clean(cell.value) for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, row))
 
 
def import_wardrobe(workbook):
    added = skipped = 0
 
    for row in sheet_rows(workbook, "Wardrobe"):
        item_name = clean(row.get("Item Name"))
        if not item_name:
            continue
 
        existing = Item.query.filter(
            db.func.lower(Item.item_name) == item_name.lower()
        ).first()
        if existing:
            skipped += 1
            continue
 
        estimated_value = clean_number(row.get("Est. Value"))
        resale = clean(row.get("Resale Willingness"))
 
        item = Item(
            item_name=item_name,
            category=clean(row.get("Category")),
            item_type=clean(row.get("Item Type")),
            colour=clean(row.get("Colour")),
            detailing=clean(row.get("Detailing")),
            style=clean(row.get("Style")),
            context=clean(row.get("Context")),
            brand=clean(row.get("Brand")),
            source_type=clean(row.get("Source Type")),
            purchase_method=clean(row.get("Purchase Method")),
            source=clean(row.get("Source")),
            country=clean(row.get("Country")),
            location=clean(row.get("Location")),
            wear_frequency=clean(row.get("Wear Frequency")),
            estimated_value=estimated_value,
            resale_willingness=resale,
            item_rating=clean(row.get("Item Rating")),
            notes=clean(row.get("Notes")),
            ownership_status=clean(row.get("Ownership Status")) or "Owned",
        )
 
        # Anything already destined for resale starts with a usable listing
        # price: the value entered when the item was first catalogued.
        if resale in SALES_RESALE:
            item.listing_price = estimated_value
 
        db.session.add(item)
        added += 1
 
    db.session.commit()
    print(f"Wardrobe: added {added}, skipped {skipped} already present.")
 
 
def import_sales(workbook):
    """Fold the Sales sheet onto the wardrobe items it describes."""
    matched = unmatched = 0
    missing_names = []
 
    for row in sheet_rows(workbook, "Sales"):
        item_name = clean(row.get("Item name"))
        if not item_name:
            continue
 
        item = Item.query.filter(
            db.func.lower(Item.item_name) == item_name.lower()
        ).first()
 
        if item is None:
            unmatched += 1
            missing_names.append(item_name)
            continue
 
        item.sale_status = clean(row.get("Status")) or "To prep"
        item.date_listed = clean_date(row.get("Date Listed"))
        item.date_sold = clean_date(row.get("Date Sold"))
        item.listing_price = clean_number(row.get("Listing Price")) or item.estimated_value
        item.sold_for = clean_number(row.get("Sold for"))
 
        # Being on the Sales sheet at all means it was up for resale.
        if item.resale_willingness not in SALES_RESALE:
            item.resale_willingness = "Sell now"
 
        matched += 1
 
    db.session.commit()
    print(f"Sales: merged {matched} onto their wardrobe items, {unmatched} unmatched.")
    for name in missing_names[:10]:
        print(f"  ! no wardrobe item named {name!r}")
 
 
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
        import_wardrobe(workbook)
        import_sales(workbook)
 
        total = Item.query.count()
        on_sales = Item.query.filter(Item.resale_willingness.in_(SALES_RESALE)).count()
        print(f"\nDone. {total} items, {on_sales} on Sales, {total - on_sales} in Wardrobe.")